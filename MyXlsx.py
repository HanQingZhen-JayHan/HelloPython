import openpyxl
from openpyxl import Workbook
fromopenpyxl.styles import PatternFill, Alignment
from openpyxl.utils.cell import get_column_letter

class MyXlsx:
    def __init(self, filename, sheetname=None):
        self.filename=filename
        self.workbook=Workbook()
        if sheetname:
            self.sheet = self.workbook.create_sheet(sheetname)
        else:
            self.sheet= self.workbook.active
            
    def create_sheet(self,sheetname=None):
        if sheetname:
            self.sheet = self.workbook.create_sheet(sheetname)
        else:
            self.sheet= self.workbook.active

    def get_sheet(self,sheetname):
        return self.workbook[sheetname]

    def write_data2column(self, list_data, column):
        self.sheet.column_dimensions[openpyxl.utils.get_column_letter(column)].width = 100
        for index in range(0,len(list_data)):
            self.sheet.cell(row=index+1, column=column, value= list_data[index].rstrip())

    def write_data2cell(self, data, row, column):
        self.sheet.cell(row, column,data)

    def read_cell(self, row, column):
        return self.sheet.cell(row, column)

    def save():
        self.workbook.active = self.sheet
        self.workbook.save(self.filename)

    def set_cell_color(self, row, column, color):
        fill=PatternFIll(start_color=color, end_color=color, fill_type='solid')
        cell_reference = get_column_letter(column)+str(row)
        self.sheet[cell_reference].fill = fill

    def write_cell_color_set(self, color_set):
        for color_param in color_set:
            self.set_cell_color(row=color_param[0], column=color_param[1], color=color_param[2])
